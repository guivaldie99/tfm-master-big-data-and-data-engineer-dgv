# weather_to_excel.py
import os
import json
import requests
import pandas as pd
from pandas import json_normalize
from datetime import datetime

API_KEY = os.getenv("MET_API_KEY")
CITY    = "Mad"
OUT_XLSX = "/Users/N218810/master/tfm/repository/meteorology-alternative/weather.xlsx"
SHEET    = "data"

URL = "http://api.weatherapi.com/v1/current.json"

def obtener_clima(ciudad, api_key):
    params = {
        "key": api_key,
        "q": ciudad,
        "aqi": "yes",
    }
    r = requests.get(URL, params=params, timeout=20)
    r.raise_for_status()
    return r.json()

def append_flat_to_excel(payload, xlsx_path, sheet_name):
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    df = json_normalize(payload)
    df.insert(0, "run_ts_utc", now)

    if not os.path.exists(xlsx_path):
        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
        return

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        startrow = ws.max_row
    else:
        startrow = 0
    wb.close()

    with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name=sheet_name,
            header=(startrow == 0),
            startrow=startrow
        )

def main():
    payload = obtener_clima(CITY, API_KEY)
    append_flat_to_excel(payload, OUT_XLSX, SHEET)
    print(f"[OK] Guardado en {OUT_XLSX}")

if __name__ == "__main__":
    main()
